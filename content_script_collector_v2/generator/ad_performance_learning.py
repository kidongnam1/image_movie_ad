from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

DEFAULT_DB = Path(__file__).resolve().parents[1] / "database" / "ad_performance.sqlite"
ANGLE_NAMES = (
    "problem_attack", "loss_aversion", "curiosity", "comparison",
    "contrarian", "discovery", "proof",
)
NUMERIC_FIELDS = (
    "impressions", "video_starts", "views_2s", "views_3s", "clicks",
    "detail_views", "purchases", "revenue", "spend",
)
TEMPLATE_FIELDS = (
    "observed_at", "campaign_id", "creative_id", "product", "category", "platform",
    "angle", "hook_text", "selling_point", "impressions", "video_starts",
    "views_2s", "views_3s", "clicks", "detail_views", "purchases", "revenue", "spend",
)

SCHEMA = """
CREATE TABLE IF NOT EXISTS performance_events (
    performance_id INTEGER PRIMARY KEY AUTOINCREMENT,
    row_fingerprint TEXT NOT NULL UNIQUE,
    observed_at TEXT,
    campaign_id TEXT,
    creative_id TEXT,
    product TEXT NOT NULL,
    category TEXT,
    platform TEXT,
    angle TEXT,
    hook_text TEXT,
    selling_point TEXT,
    impressions INTEGER NOT NULL DEFAULT 0,
    video_starts INTEGER NOT NULL DEFAULT 0,
    views_2s INTEGER NOT NULL DEFAULT 0,
    views_3s INTEGER NOT NULL DEFAULT 0,
    clicks INTEGER NOT NULL DEFAULT 0,
    detail_views INTEGER NOT NULL DEFAULT 0,
    purchases INTEGER NOT NULL DEFAULT 0,
    revenue REAL NOT NULL DEFAULT 0,
    spend REAL NOT NULL DEFAULT 0,
    source_file TEXT,
    imported_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_perf_category_angle ON performance_events(category, angle);
CREATE INDEX IF NOT EXISTS idx_perf_creative ON performance_events(creative_id);
CREATE INDEX IF NOT EXISTS idx_perf_product ON performance_events(product);
"""


def connect(db_path: str | Path | None = None) -> sqlite3.Connection:
    path = Path(db_path or DEFAULT_DB)
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)
    return conn


def _to_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    return int(float(str(value).replace(",", "").strip()))


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(str(value).replace(",", "").strip())


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {k: str(row.get(k, "") or "").strip() for k in TEMPLATE_FIELDS}
    out["product"] = out["product"].strip()
    if not out["product"]:
        raise ValueError("product is required")
    angle = out["angle"].strip().lower()
    if angle and angle not in ANGLE_NAMES:
        raise ValueError(f"unsupported angle: {angle}")
    out["angle"] = angle
    for k in NUMERIC_FIELDS:
        out[k] = _to_float(row.get(k)) if k in {"revenue", "spend"} else _to_int(row.get(k))
        if out[k] < 0:
            raise ValueError(f"{k} must be >= 0")
    imp = out["impressions"]
    for k in ("video_starts", "views_2s", "views_3s", "clicks", "detail_views", "purchases"):
        if imp and out[k] > imp:
            raise ValueError(f"{k} cannot exceed impressions")
    if out["views_3s"] > out["views_2s"] and out["views_2s"] > 0:
        raise ValueError("views_3s cannot exceed views_2s")
    return out


def row_fingerprint(row: dict[str, Any]) -> str:
    payload = "|".join(str(row.get(k, "")) for k in TEMPLATE_FIELDS)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _rows_from_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _rows_from_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        data = data.get("rows", data.get("performance", [data]))
    if not isinstance(data, list):
        raise ValueError("JSON must contain a list of performance rows")
    return [dict(x) for x in data]


def _rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX import requires openpyxl. Run: pip install openpyxl") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = ws.iter_rows(values_only=True)
    headers = [str(x or "").strip() for x in next(values)]
    return [{headers[i]: v for i, v in enumerate(row) if i < len(headers)} for row in values]


def read_rows(path: str | Path) -> list[dict[str, Any]]:
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".csv":
        return _rows_from_csv(p)
    if ext == ".json":
        return _rows_from_json(p)
    if ext in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx(p)
    raise ValueError("supported performance files: .csv, .json, .xlsx, .xlsm")


def import_file(path: str | Path, db_path: str | Path | None = None) -> dict[str, Any]:
    p = Path(path)
    rows = read_rows(p)
    conn = connect(db_path)
    inserted = skipped = errors = 0
    error_rows: list[dict[str, Any]] = []
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    sql = """
        INSERT OR IGNORE INTO performance_events (
            row_fingerprint, observed_at, campaign_id, creative_id, product, category, platform,
            angle, hook_text, selling_point, impressions, video_starts, views_2s, views_3s,
            clicks, detail_views, purchases, revenue, spend, source_file, imported_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """
    for idx, raw in enumerate(rows, start=2):
        try:
            r = normalize_row(raw)
            fp = row_fingerprint(r)
            cur = conn.execute(sql, (
                fp, r["observed_at"], r["campaign_id"], r["creative_id"], r["product"],
                r["category"], r["platform"], r["angle"], r["hook_text"], r["selling_point"],
                r["impressions"], r["video_starts"], r["views_2s"], r["views_3s"],
                r["clicks"], r["detail_views"], r["purchases"], r["revenue"], r["spend"],
                p.name, now,
            ))
            inserted += 1 if cur.rowcount else 0
            skipped += 0 if cur.rowcount else 1
        except Exception as exc:
            errors += 1
            error_rows.append({"row": idx, "error": str(exc)})
    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM performance_events").fetchone()[0]
    conn.close()
    return {"source": str(p), "inserted": inserted, "duplicates_skipped": skipped, "errors": errors,
            "error_rows": error_rows[:20], "database_rows": total}


def create_template(path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerow(TEMPLATE_FIELDS)
    return p


def _aggregate(rows: Iterable[sqlite3.Row | dict[str, Any]]) -> dict[str, float]:
    s = {k: 0.0 for k in NUMERIC_FIELDS}
    count = 0
    for row in rows:
        count += 1
        for k in NUMERIC_FIELDS:
            s[k] += float(row[k] or 0)
    s["rows"] = float(count)
    return s


def _rate(num: float, den: float) -> float:
    return num / den if den > 0 else 0.0


def _metrics(agg: dict[str, float]) -> dict[str, float]:
    imp = agg["impressions"]
    clicks = agg["clicks"]
    spend = agg["spend"]
    return {
        "impressions": imp,
        "rows": agg.get("rows", 0.0),
        "retention_2s": _rate(agg["views_2s"], imp),
        "retention_3s": _rate(agg["views_3s"], imp),
        "ctr": _rate(clicks, imp),
        "purchase_cvr": _rate(agg["purchases"], clicks),
        "purchase_rate": _rate(agg["purchases"], imp),
        "roas": _rate(agg["revenue"], spend),
    }


def _smooth_rate(success: float, den: float, prior: float, strength: float) -> float:
    if den <= 0:
        return prior
    return (success + prior * strength) / (den + strength)


def _lift(value: float, baseline: float, clip: float = 0.60) -> float:
    if baseline <= 0:
        return 0.0
    return max(-clip, min(clip, value / baseline - 1.0))


def _angle_adjustment(angle_agg: dict[str, float], baseline: dict[str, float]) -> tuple[float, dict[str, float]]:
    imp = angle_agg["impressions"]
    clicks = angle_agg["clicks"]
    r2 = _smooth_rate(angle_agg["views_2s"], imp, baseline["retention_2s"], 600)
    r3 = _smooth_rate(angle_agg["views_3s"], imp, baseline["retention_3s"], 600)
    ctr = _smooth_rate(clicks, imp, baseline["ctr"], 800)
    cvr = _smooth_rate(angle_agg["purchases"], clicks, baseline["purchase_cvr"], 80)
    roas = _metrics(angle_agg)["roas"] if angle_agg["spend"] > 0 else baseline["roas"]
    weighted_lift = (
        _lift(r2, baseline["retention_2s"]) * .25 +
        _lift(r3, baseline["retention_3s"]) * .15 +
        _lift(ctr, baseline["ctr"]) * .25 +
        _lift(cvr, baseline["purchase_cvr"]) * .25 +
        _lift(roas, baseline["roas"]) * .10
    )
    confidence = 1.0 - math.exp(-imp / 2000.0)
    if imp < 200:
        confidence *= imp / 200.0
    adjustment = max(-6.0, min(6.0, weighted_lift * 12.0 * confidence))
    return round(adjustment, 2), {
        "retention_2s": round(r2, 6), "retention_3s": round(r3, 6),
        "ctr": round(ctr, 6), "purchase_cvr": round(cvr, 6), "roas": round(roas, 4),
        "confidence": round(confidence, 4), "impressions": int(imp),
    }


def build_learning_profile(category: str = "", db_path: str | Path | None = None) -> dict[str, Any]:
    conn = connect(db_path)
    all_rows = conn.execute("SELECT * FROM performance_events").fetchall()
    if not all_rows:
        conn.close()
        return {
            "active": False, "reason": "no performance data", "total_rows": 0,
            "total_impressions": 0, "category": category, "baselines": {},
            "angle_adjustments": {a: 0.0 for a in ANGLE_NAMES}, "angle_details": {},
        }
    all_agg = _aggregate(all_rows)
    baseline = _metrics(all_agg)
    adjustments: dict[str, float] = {}
    details: dict[str, Any] = {}
    for angle in ANGLE_NAMES:
        global_rows = [r for r in all_rows if (r["angle"] or "") == angle]
        cat_rows = [r for r in global_rows if category and (r["category"] or "") == category]
        global_agg = _aggregate(global_rows)
        global_adj, global_detail = _angle_adjustment(global_agg, baseline) if global_rows else (0.0, {})
        if cat_rows:
            cat_agg = _aggregate(cat_rows)
            cat_adj, cat_detail = _angle_adjustment(cat_agg, baseline)
            cat_conf = 1.0 - math.exp(-cat_agg["impressions"] / 1500.0)
            adj = cat_adj * cat_conf + global_adj * (1.0 - cat_conf) * 0.5
            details[angle] = {"category": cat_detail, "global": global_detail, "category_blend": round(cat_conf, 4)}
        else:
            adj = global_adj * 0.5
            details[angle] = {"category": None, "global": global_detail, "category_blend": 0.0}
        adjustments[angle] = round(max(-6.0, min(6.0, adj)), 2)
    conn.close()
    return {
        "active": True, "reason": "performance data available", "category": category,
        "total_rows": int(all_agg["rows"]), "total_impressions": int(all_agg["impressions"]),
        "baselines": {k: round(v, 6) if isinstance(v, float) else v for k, v in baseline.items()},
        "angle_adjustments": adjustments, "angle_details": details,
        "weights": {"retention_2s": .25, "retention_3s": .15, "ctr": .25, "purchase_cvr": .25, "roas": .10},
        "max_adjustment": 6.0,
    }


def creative_id(product: str, category: str, angle: str, hook_text: str) -> str:
    raw = f"v26|{product}|{category}|{angle}|{hook_text}"
    return "CR-" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12].upper()


def report(db_path: str | Path | None = None, category: str = "") -> dict[str, Any]:
    return build_learning_profile(category=category, db_path=db_path)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="V2.6 ad performance learning store")
    sub = ap.add_subparsers(dest="cmd", required=True)
    p = sub.add_parser("init-db"); p.add_argument("--db", default=str(DEFAULT_DB))
    p = sub.add_parser("template"); p.add_argument("--out", default="ad_performance_template.csv")
    p = sub.add_parser("import"); p.add_argument("file"); p.add_argument("--db", default=str(DEFAULT_DB))
    p = sub.add_parser("report"); p.add_argument("--db", default=str(DEFAULT_DB)); p.add_argument("--category", default="")
    return ap.parse_args()


def main() -> int:
    a = parse_args()
    if a.cmd == "init-db":
        c = connect(a.db); c.close(); print("DB_READY=", a.db); return 0
    if a.cmd == "template":
        print("TEMPLATE=", create_template(a.out)); return 0
    if a.cmd == "import":
        print(json.dumps(import_file(a.file, a.db), ensure_ascii=False, indent=2)); return 0
    if a.cmd == "report":
        print(json.dumps(report(a.db, a.category), ensure_ascii=False, indent=2)); return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
