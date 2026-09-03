from __future__ import annotations
import argparse, csv, json, re, tempfile
from pathlib import Path
from typing import Any

try:
    from . import performance_store_v26 as store
    from . import performance_registry_v27 as registry
except ImportError:
    import performance_store_v26 as store  # type: ignore
    import performance_registry_v27 as registry  # type: ignore

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = ROOT / "database" / "ad_performance.sqlite"
ALIASES = {
    "observed_at": ["date","day","reporting starts","reporting start","날짜","일자","기준일","기간"],
    "campaign_id": ["campaign id","campaign_id","캠페인 id","캠페인id","campaign","campaign name","캠페인","캠페인명"],
    "creative_id": ["creative id","creative_id","ad id","ad_id","광고 id","광고id","소재 id","소재id","ad name","광고명","소재명"],
    "product": ["product","product name","상품","상품명","item","item name"],
    "category": ["category","카테고리"],
    "angle": ["angle","creative angle","광고각도","앵글"],
    "hook_text": ["hook","hook text","hook_text","후킹","후킹문구","광고문구","소재문구"],
    "selling_point": ["selling point","selling_point","핵심특징","강조특징","소구점"],
    "impressions": ["impressions","impression","노출","노출수","노출수(회)"],
    "video_starts": ["video starts","video plays","video views","동영상 재생","동영상조회","재생수","영상조회수"],
    "views_2s": ["2-second continuous video plays","2 second continuous video plays","2-second video views","2 second video views","2s video views","2초 동영상 조회","2초 조회","2초재생"],
    "views_3s": ["3-second video plays","3 second video plays","3-second video views","3 second video views","3s video views","3초 동영상 조회","3초 조회","3초재생"],
    "clicks": ["link clicks","outbound clicks","clicks","click","클릭","클릭수"],
    "detail_views": ["landing page views","landing page view","destination page views","detail views","상세페이지 조회","상세페이지 유입","랜딩페이지 조회","랜딩페이지뷰"],
    "purchases": ["purchases","purchase","orders","order","conversions","구매","구매수","주문","주문수","전환","전환수"],
    "revenue": ["purchase conversion value","conversion value","revenue","sales","sales amount","매출","매출액","전환매출","전환매출액","구매금액"],
    "spend": ["amount spent","spend","cost","ad spend","광고비","비용","총비용","소진액"],
}
PLATFORM_HINTS = {
    "meta": ["reporting starts","amount spent","2-second continuous video plays","landing page views","facebook","instagram"],
    "tiktok": ["2-second video views","6-second video views","tiktok","video views at 25%","cost per 1000 people reached"],
    "naver": ["네이버","파워링크","쇼핑검색","전환매출액","총비용","키워드"],
    "coupang": ["쿠팡","coupang","주문수","광고비","판매수익","로켓그로스"],
}

def norm(s: Any) -> str:
    return re.sub(r"[^0-9a-z가-힣]+", "", str(s or "").strip().lower())
ALIAS_NORM = {k: [norm(x) for x in vals] for k, vals in ALIASES.items()}

def read_tabular(path: str | Path) -> list[dict[str, Any]]:
    p = Path(path); ext = p.suffix.lower()
    if ext == ".csv":
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                with p.open("r", encoding=enc, newline="") as f: return list(csv.DictReader(f))
            except UnicodeDecodeError: continue
        raise UnicodeError("CSV encoding not supported: tried UTF-8 BOM/UTF-8, CP949, EUC-KR")
    if ext == ".json":
        d = json.loads(p.read_text(encoding="utf-8-sig")); d = d.get("rows", d.get("data", [d])) if isinstance(d, dict) else d
        if not isinstance(d, list): raise ValueError("JSON must contain a list of rows")
        return [dict(x) for x in d]
    if ext in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        wb=load_workbook(p, read_only=True, data_only=True); ws=wb[wb.sheetnames[0]]; it=ws.iter_rows(values_only=True)
        try: headers=[str(x or "").strip() for x in next(it)]
        except StopIteration: return []
        return [{headers[i]:v for i,v in enumerate(row) if i < len(headers)} for row in it]
    raise ValueError("supported source files: .csv, .json, .xlsx, .xlsm")

def detect_platform(headers: list[str]) -> str:
    hn = {norm(h) for h in headers}; scores = {}
    for platform, hints in PLATFORM_HINTS.items(): scores[platform] = sum(norm(h) in hn or any(norm(h) in x for x in hn) for h in hints)
    best = max(scores, key=scores.get) if scores else "generic"
    return best if scores.get(best, 0) > 0 else "generic"

def match_headers(headers: list[str]) -> tuple[dict[str, str], list[str]]:
    nmap = {norm(h): h for h in headers}; mapping: dict[str, str] = {}; used: set[str] = set()
    for canonical, aliases in ALIAS_NORM.items():
        for alias in aliases:
            if alias in nmap: mapping[canonical]=nmap[alias]; used.add(nmap[alias]); break
            hit = next((orig for nh, orig in nmap.items() if alias and (alias in nh or nh in alias)), None)
            if hit: mapping[canonical]=hit; used.add(hit); break
    return mapping, [h for h in headers if h not in used]

def val(row: dict[str, Any], mapping: dict[str, str], key: str, default: Any = "") -> Any:
    src = mapping.get(key); return row.get(src, default) if src else default

def normalize_source(path: str | Path, platform: str = "auto", db_path: str | Path = DEFAULT_DB, default_product: str = "", default_category: str = "") -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows = read_tabular(path); headers = list(rows[0].keys()) if rows else []; detected = detect_platform(headers) if platform in {"", "auto"} else platform.lower(); mapping, unmapped = match_headers(headers)
    conn = store.connect(db_path); registry.ensure_registry(conn); normalized=[]; enriched=0; missing_product=0
    for raw in rows:
        cr = registry.extract_creative_id(*raw.values()); meta = registry.lookup(conn, cr) if cr else None
        if meta: enriched += 1
        product = str(val(raw,mapping,"product","") or (meta or {}).get("product") or default_product).strip()
        category = str(val(raw,mapping,"category","") or (meta or {}).get("category") or default_category).strip()
        angle = str(val(raw,mapping,"angle","") or (meta or {}).get("angle") or "").strip().lower()
        hook = str(val(raw,mapping,"hook_text","") or (meta or {}).get("hook_text") or "").strip(); selling = str(val(raw,mapping,"selling_point","") or (meta or {}).get("selling_point") or "").strip(); cid = cr or str(val(raw,mapping,"creative_id","") or "").strip()
        if not product: missing_product += 1
        normalized.append({"observed_at":val(raw,mapping,"observed_at",""),"campaign_id":val(raw,mapping,"campaign_id",""),"creative_id":cid,"product":product,"category":category,"platform":detected,"angle":angle,"hook_text":hook,"selling_point":selling,"impressions":val(raw,mapping,"impressions",0),"video_starts":val(raw,mapping,"video_starts",0),"views_2s":val(raw,mapping,"views_2s",0),"views_3s":val(raw,mapping,"views_3s",0),"clicks":val(raw,mapping,"clicks",0),"detail_views":val(raw,mapping,"detail_views",0),"purchases":val(raw,mapping,"purchases",0),"revenue":val(raw,mapping,"revenue",0),"spend":val(raw,mapping,"spend",0)})
    conn.close(); report={"source":str(path),"platform":detected,"rows":len(rows),"mapped_headers":mapping,"unmapped_headers":unmapped,"registry_enriched_rows":enriched,"missing_product_rows":missing_product,"warnings":(["일부 행의 product를 복원하지 못했습니다. 광고명에 CR-... ID를 넣거나 --default-product를 지정하세요."] if missing_product else [])}; return normalized, report

def write_normalized(rows: list[dict[str, Any]], out: str | Path) -> Path:
    p=Path(out); p.parent.mkdir(parents=True,exist_ok=True)
    with p.open("w",encoding="utf-8-sig",newline="") as f: w=csv.DictWriter(f,fieldnames=store.TEMPLATE_FIELDS);w.writeheader();w.writerows(rows)
    return p

def import_platform(path: str|Path, platform="auto", db_path: str|Path=DEFAULT_DB, default_product="", default_category="", normalized_out: str|Path|None=None) -> dict[str,Any]:
    rows, rep = normalize_source(path,platform,db_path,default_product,default_category); rows=[r for r in rows if str(r.get("product","")).strip()]; out = Path(normalized_out) if normalized_out else Path(tempfile.gettempdir())/"v27_normalized_performance.csv"; write_normalized(rows,out); result = store.import_file(out,db_path); return {"adapter":rep,"import":result,"normalized_file":str(out)}

def parse_args():
    p=argparse.ArgumentParser(description="V2.7 platform performance adapter");p.add_argument("file");p.add_argument("--platform",choices=("auto","meta","tiktok","naver","coupang","generic"),default="auto");p.add_argument("--db",default=str(DEFAULT_DB));p.add_argument("--default-product",default="");p.add_argument("--default-category",default="");p.add_argument("--normalized-out",default="");return p.parse_args()
def main():
    a=parse_args();r=import_platform(a.file,a.platform,a.db,a.default_product,a.default_category,a.normalized_out or None);print(json.dumps(r,ensure_ascii=False,indent=2));return 0
if __name__=="__main__":raise SystemExit(main())
